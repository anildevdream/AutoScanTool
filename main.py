import time

from selenium import webdriver
import ProviderList
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

from openpyxl import Workbook
from openpyxl.styles import Font

# Configure webdriver options
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("webdriver.chrome.driver=[D:/ChromeDriver/chromedriver.exe]")

driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
driver.delete_all_cookies()
driver.get("https://scan.doctor.com/2two5%20co-branded")
print("Logged in to scan.doctor.com")

wb = Workbook()
ws = wb.active
excel_headers = ["PROVIDER_NAME", "LISTINGS_ACCURACY", "PROVIDER_LISTINGS-GOOGLE", "HEALTHGRADE", "WEB MD", "VITALS",
                 "SHARECARE", "WELLNESS", "RATEMDS", "DOCTOR", "NPPES", "FACILITY LISTINGS-GOOGLE", "BING",
                 "DOCTOR.COM"]
ws.append(excel_headers)
for cell in ws[1]:
    cell.font = Font(bold=True)
cell_starting_count = 2
for link in ProviderList.links_to_open:
    # Open a new window
    driver.execute_script("window.open('%s', '_blank')" % link)
    driver.switch_to.window(driver.window_handles[-1])
    wait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Scan Now']"))).click()
    wait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "scan-header-results")))
    provider_name = driver.find_element(By.XPATH,
                                        f"//section[@class='provider__container ng-scope']//div[@class='detailed-report ng-scope']//section[@class='providerSummary']//div[@class='providerColumn nameColumn']//span[@data-qa='name']").text
    print(f"Scanning done for provider:{provider_name}")
    listing_accuracy = driver.find_element(By.XPATH,
                                           f"//section[@class='providerSummary']//div[@class='providerColumn scoreColumn']//div[@data-qa='accuracy']").text
    print(f"Listings Accuracy of provider: {provider_name}: {listing_accuracy}")
    source_list = ["googleProvider", "healthgrades", "webmd", "vitals", "sharecare", "wellness", "ratemds", "doctor",
                   "NPPES"]
    accuracy_elements = []
    for source in source_list:
        try:
            path = driver.find_element(By.XPATH,
                                       f"//div[@data-qa-listing='{source}']//li//span[not(contains(@class,'ng-hide')) and contains(@class,'detailed-report__status') and (contains(@class,'status-error') or contains(@class,'status-success'))]")
            accuracy_element = path.get_attribute("innerHTML")
            accuracy_elements.append(accuracy_element)
            print(f"{source}:{accuracy_element}")
        except NoSuchElementException:
            print("NO SOURCE FOUND")
    facility_list = ["google", "bingPractice", "doctorPractice"]
    facility_list_elements = []
    for facility_source in facility_list:
        try:
            facility_path = driver.find_element(By.XPATH,
                                                f"//section[@class='local-search']//ul//li[@data-qa-listing='{facility_source}']//span[not(contains(@class,'ng-hide')) and contains(@class,'detailed-report__status') and (contains(@class,'status-error') or contains(@class,'status-success'))]")
            facility_element = facility_path.get_attribute("innerHTML")
            facility_list_elements.append(facility_element)
            print(f"{facility_source}:{facility_element}")
        except NoSuchElementException:
            print("NO SOURCE FOUND")

    # Write Scan results to Excel
    ws[f'A{cell_starting_count}'] = provider_name
    ws[f'B{cell_starting_count}'] = listing_accuracy
    ws[f'C{cell_starting_count}'] = accuracy_elements[0]
    ws[f'D{cell_starting_count}'] = accuracy_elements[1]
    ws[f'E{cell_starting_count}'] = accuracy_elements[2]
    ws[f'F{cell_starting_count}'] = accuracy_elements[3]
    ws[f'G{cell_starting_count}'] = accuracy_elements[4]
    ws[f'H{cell_starting_count}'] = accuracy_elements[5]
    ws[f'I{cell_starting_count}'] = accuracy_elements[6]
    ws[f'J{cell_starting_count}'] = accuracy_elements[7]
    ws[f'K{cell_starting_count}'] = accuracy_elements[8]
    ws[f'L{cell_starting_count}'] = facility_list_elements[0]
    ws[f'M{cell_starting_count}'] = facility_list_elements[1]
    ws[f'N{cell_starting_count}'] = facility_list_elements[2]
    cell_starting_count = cell_starting_count + 1
    time.sleep(3)

    for row in ws.iter_rows(min_row=2, max_row=13, min_col=3, max_col=14):
        for cell in row:
            if cell.value == "Good":
                cell.font = Font(color="008000")  # Green
            else:
                cell.font = Font(color="FF0000")  # Red
wb.save("C:/Users/kumara/Desktop/Codes/listing.xlsx")
wb.close()





